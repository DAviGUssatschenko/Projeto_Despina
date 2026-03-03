"""
fetch_climate.py
----------------
Extrai dados climáticos do banco PostgreSQL para os casos de sinistro
identificados no Excel. Usa estratégia de proximidade espacial (N pontos
mais próximos ao centróide de cada polígono) dado que a resolução da
grade (~10 km) é maior que os talhões agrícolas (<500 m).

Uso:
    python src/seca/data/fetch_climate.py

Saída:
    data/processed/seca_all_data.csv
    data/processed/seca_summary.csv
"""

import re
import time
import warnings
import numpy as np
import pandas as pd
import openpyxl

from datetime import datetime
from pathlib import Path
from shapely import wkt as shapely_wkt
from sqlalchemy import create_engine, text
from tqdm import tqdm

warnings.filterwarnings("ignore")

# =============================================================================
# CONFIGURAÇÃO
# =============================================================================

DB_CONFIG = {
    "host":     "localhost",
    "port":     5432,
    "database": "nome_do_banco",
    "user":     "usuario",
    "password": "senha",
}

TABLE_POINTS = "points_coordinates"
TABLE_DATA   = "nome_da_tabela_climatica"

EXCEL_PATH  = Path("data/raw/DADOS_TASK_BANDAS_DS_TEAM.xlsx")
OUTPUT_DIR  = Path("data/processed")

N_NEAREST      = 1      # pontos mais próximos por polígono
MAX_DIST_DEG   = 0.18   # ~20 km em graus

# =============================================================================
# UTILITÁRIOS
# =============================================================================

def parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, str):
        val = val.strip()
        for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                continue
    return None


def fix_wkt(raw):
    parts = re.split(r"\n\s*\n", raw)
    parts = [p.strip() for p in parts if p.strip()]
    fixed = []
    for p in parts:
        p = re.sub(r"([0-9])(-)([0-9])", r"\1 \2\3", p)
        p = re.sub(r",(?=[^\s])", ", ", p)
        fixed.append(p)
    if len(fixed) == 1:
        return fixed[0]
    inner = ", ".join(
        re.sub(r"^POLYGON\s*", "", f, flags=re.IGNORECASE).strip()
        for f in fixed
    )
    return f"MULTIPOLYGON ({inner})"


def read_seca_records(excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    SKIP = {None, "-", "id", "geometry", "WHEAT", "OTHERS",
            "TRIGO", "MILHO", "SOJA", "CORN", "SOY"}
    records = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            if not row or len(row) < 6:
                continue
            id_val, geometry, _, start_raw, end_raw, complaint = row[:6]
            if id_val in SKIP or not isinstance(id_val, (int, float)):
                continue
            if geometry is None or str(geometry).strip() in SKIP:
                continue
            if str(complaint).strip().lower() != "seca":
                continue
            s = parse_date(start_raw)
            e = parse_date(end_raw)
            if s is None or e is None:
                continue
            if e < s:
                s, e = e, s
            try:
                geom = shapely_wkt.loads(fix_wkt(str(geometry).strip()))
            except Exception:
                continue
            records.append({
                "sheet": sheet_name, "case_id": int(id_val),
                "complaint": str(complaint).strip(),
                "geometry": geom, "start_date": s, "end_date": e,
            })
    df = pd.DataFrame(records).drop_duplicates(
        subset=["case_id", "start_date", "end_date"])
    print(f"[INFO] {len(df)} registros únicos de 'Seca' encontrados.")
    return df


def load_all_points(engine):
    q = f"SELECT point_id, latitude, longitude FROM {TABLE_POINTS}"
    df = pd.read_sql(q, engine)
    return df


def find_nearest_points(geom, points_df, n=1, max_dist=0.18):
    centroid = geom.centroid
    cx, cy = centroid.x, centroid.y
    cand = points_df[
        (points_df["longitude"].between(cx - max_dist, cx + max_dist)) &
        (points_df["latitude"].between(cy - max_dist, cy + max_dist))
    ].copy()
    if cand.empty:
        return []
    dlat = (cand["latitude"] - cy).values
    dlon = (cand["longitude"] - cx).values
    cand["dist"] = np.sqrt(dlat**2 + dlon**2)
    nearest = cand.nsmallest(n, "dist")
    return list(nearest["point_id"].values), list(nearest["dist"].values)


def query_data(engine, point_ids, start_date, end_date):
    ids = ", ".join(str(int(p)) for p in point_ids)
    q = f"""
        SELECT *
        FROM   {TABLE_DATA}
        WHERE  point_id IN ({ids})
          AND  date >= '{start_date}'
          AND  date <= '{end_date}'
        ORDER  BY date, point_id
    """
    return pd.read_sql(q, engine)


# =============================================================================
# MAIN
# =============================================================================

def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    cases_df = read_seca_records(EXCEL_PATH)
    if cases_df.empty:
        print("Nenhum caso encontrado.")
        return

    url = (f"postgresql+psycopg2://{DB_CONFIG['user']}:{DB_CONFIG['password']}"
           f"@{DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['database']}")
    engine = create_engine(url)
    print("Conexão com banco estabelecida.")

    points_df = load_all_points(engine)
    print(f"[INFO] {len(points_df):,} pontos carregados.")

    all_data, summary = [], []

    for _, row in tqdm(cases_df.iterrows(), total=len(cases_df), desc="Casos"):
        case_id = row["case_id"]
        result  = find_nearest_points(row["geometry"], points_df, N_NEAREST, MAX_DIST_DEG)
        if not result or not result[0]:
            continue
        point_ids, dists = result

        df = query_data(engine, point_ids, row["start_date"], row["end_date"])
        if df.empty:
            continue

        df["case_id"]       = case_id
        df["sheet"]         = row["sheet"]
        df["start_case"]    = row["start_date"]
        df["end_case"]      = row["end_date"]
        df["nearest_dist_km"] = round(dists[0] * 111.32, 2)
        all_data.append(df)

        summary.append({
            "case_id":          case_id,
            "sheet":            row["sheet"],
            "start_date":       row["start_date"],
            "end_date":         row["end_date"],
            "nearest_point_id": point_ids[0],
            "nearest_dist_km":  round(dists[0] * 111.32, 2),
            "rows_retrieved":   len(df),
        })

    if all_data:
        pd.concat(all_data, ignore_index=True).to_csv(
            OUTPUT_DIR / "seca_all_data.csv", index=False)
    pd.DataFrame(summary).to_csv(
        OUTPUT_DIR / "seca_summary.csv", index=False)

    print(f"\n[OK] seca_all_data.csv e seca_summary.csv salvos em {OUTPUT_DIR}")


if __name__ == "__main__":
    main()

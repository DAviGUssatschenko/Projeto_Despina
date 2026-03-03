"""
fetch_copernicus.py
-------------------
Baixa índices espectrais Sentinel-2 (L2A) do Copernicus Data Space para
todos os casos do Excel. Calcula 6 índices especializados por imagem:
NBR, NDRE, MSI, GNDVI, SAVI, NDDI.

Uso:
    python src/seca/data/fetch_copernicus.py

Saída:
    data/processed/output_indices/
        case_{id}_{complaint}/
            {id}_{data}.tif          <- GeoTIFF 6 bandas
            {id}_stats.csv           <- estatísticas por índice/data
            {id}_timeseries.csv      <- série temporal pivotada
        timeseries_global.csv
        resumo_geral.csv
        media_por_complaint.csv
"""

import re, json, time, warnings
import numpy as np
import pandas as pd
import geopandas as gpd
import requests
import rasterio
import openpyxl

from datetime import datetime
from pathlib import Path
from shapely import wkt as shapely_wkt
from tqdm import tqdm

warnings.filterwarnings("ignore")

# =============================================================================
# CONFIGURAÇÃO
# =============================================================================

CLIENT_ID     = "sh-xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxxxxxx"   # OAuth client S-Hub
CLIENT_SECRET = "seu_secret_aqui"

EXCEL_PATH = Path("data/raw/DADOS_TASK_BANDAS_DS_TEAM.xlsx")
OUTPUT_DIR = Path("data/processed/output_indices")

MAX_CLOUD  = 20    # % máximo de cobertura de nuvens
MAX_SCENES = 20    # máximo de imagens por caso
IMG_SIZE   = 512   # resolução em pixels

INDEX_NAMES = ["NBR", "NDRE", "MSI", "GNDVI", "SAVI", "NDDI"]

INDEX_BOUNDS = {
    "NBR":   (-1.0,  1.0),
    "NDRE":  (-1.0,  1.0),
    "MSI":   ( 0.0, 20.0),
    "GNDVI": (-1.0,  1.0),
    "SAVI":  (-1.5,  1.5),
    "NDDI":  (-1.0,  1.0),
}

PERCENTILES = [5, 10, 25, 50, 75, 90, 95]

# =============================================================================
# AUTENTICAÇÃO
# =============================================================================

def get_token():
    url = ("https://identity.dataspace.copernicus.eu"
           "/auth/realms/CDSE/protocol/openid-connect/token")
    resp = requests.post(url, data={
        "grant_type":    "client_credentials",
        "client_id":     CLIENT_ID,
        "client_secret": CLIENT_SECRET,
    }, timeout=30)
    if resp.status_code != 200:
        raise Exception(f"Auth falhou: {resp.text}")
    print("[OK] Token obtido.")
    return resp.json()["access_token"]

def get_headers(token):
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

# =============================================================================
# RETRY
# =============================================================================

def request_with_retry(method, url, max_tries=4, wait=5, **kwargs):
    for attempt in range(max_tries):
        try:
            return getattr(requests, method)(url, timeout=120, **kwargs)
        except (requests.exceptions.ConnectionError,
                requests.exceptions.Timeout) as e:
            if attempt == max_tries - 1:
                raise
            sleep = wait * (2 ** attempt)
            tqdm.write(f"  [RETRY {attempt+1}/{max_tries}] aguardando {sleep}s")
            time.sleep(sleep)

# =============================================================================
# LEITURA DO EXCEL
# =============================================================================

def parse_date(val):
    if val is None: return None
    if isinstance(val, datetime): return val.date()
    if isinstance(val, str):
        val = val.strip()
        for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
            try: return datetime.strptime(val, fmt).date()
            except ValueError: continue
        parts = re.split(r"[/\-]", val)
        if len(parts) == 3:
            try:
                d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
                if y < 100: y += 2000
                return datetime(y, m, d).date()
            except: pass
    return None

def fix_wkt(raw):
    parts = re.split(r"\n\s*\n", raw)
    parts = [p.strip() for p in parts if p.strip()]
    fixed = []
    for p in parts:
        p = re.sub(r"([0-9])(-)([0-9])", r"\1 \2\3", p)
        p = re.sub(r",(?=[^\s])", ", ", p)
        fixed.append(p)
    if len(fixed) == 1:
        return fixed[0]
    inner = ", ".join(
        re.sub(r"^POLYGON\s*", "", f, flags=re.IGNORECASE).strip() for f in fixed)
    return f"MULTIPOLYGON ({inner})"

def read_all_records(excel_path):
    wb   = openpyxl.load_workbook(excel_path, data_only=True)
    SKIP = {None, "-", "id", "geometry", "WHEAT", "OTHERS",
            "TRIGO", "MILHO", "SOJA", "CORN", "SOY"}
    records = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            if not row or len(row) < 6: continue
            id_val, geometry, _, start_raw, end_raw, complaint = row[:6]
            if id_val in SKIP or not isinstance(id_val, (int, float)): continue
            if geometry is None or str(geometry).strip() in SKIP: continue
            if complaint is None or str(complaint).strip() in SKIP: continue
            s = parse_date(start_raw); e = parse_date(end_raw)
            if s is None or e is None: continue
            if e < s: s, e = e, s
            try: geom = shapely_wkt.loads(fix_wkt(str(geometry).strip()))
            except: continue
            records.append({"sheet": sheet_name, "case_id": int(id_val),
                             "complaint": str(complaint).strip(),
                             "geometry": geom, "start_date": s, "end_date": e})
    df = pd.DataFrame(records).drop_duplicates(
        subset=["case_id", "start_date", "end_date", "complaint"])
    print(f"[INFO] {len(df)} casos válidos no Excel.")
    return df

# =============================================================================
# GEOMETRIA → API
# =============================================================================

def geom_to_api(geom):
    safe = geom.convex_hull
    gdf  = gpd.GeoDataFrame(geometry=[safe], crs="EPSG:4326")
    aoi  = json.loads(gdf.to_json())["features"][0]["geometry"]
    return aoi, list(safe.bounds)

# =============================================================================
# STAC
# =============================================================================

def search_scenes(bbox, start, end, headers, max_cloud=20, limit=20):
    url     = "https://catalogue.dataspace.copernicus.eu/stac/search"
    payload = {"collections": ["sentinel-2-l2a"], "bbox": bbox,
                "datetime": f"{start}T00:00:00Z/{end}T23:59:59Z",
                "limit": limit, "query": {"eo:cloud_cover": {"lte": max_cloud}}}
    try:
        resp = request_with_retry("post", url, json=payload, headers=headers)
        if resp.status_code != 200: return []
        feats = resp.json().get("features", [])
        feats.sort(key=lambda x: x["properties"]["datetime"])
        return feats
    except Exception as e:
        tqdm.write(f"  [STAC ERROR] {e}"); return []

# =============================================================================
# EVALSCRIPT — 6 índices
# =============================================================================

EVALSCRIPT = """
//VERSION=3
function setup() {
  return {
    input:  [{ bands: ["B03","B04","B05","B08","B8A","B11","B12"], units: "REFLECTANCE" }],
    output: { bands: 6, sampleType: "FLOAT32" }
  };
}
function evaluatePixel(s) {
  var e   = 1e-10;
  var nbr  = (s.B08 - s.B12) / (s.B08 + s.B12 + e);
  var ndre = (s.B8A - s.B05) / (s.B8A + s.B05 + e);
  var msi  = s.B11 / (s.B08 + e);
  var gndvi= (s.B08 - s.B03) / (s.B08 + s.B03 + e);
  var savi = 1.5 * (s.B08 - s.B04) / (s.B08 + s.B04 + 0.5 + e);
  var ndvi = (s.B08 - s.B04) / (s.B08 + s.B04 + e);
  var ndwi = (s.B08 - s.B11) / (s.B08 + s.B11 + e);
  var nddi = (ndvi - ndwi) / (Math.abs(ndvi + ndwi) + e);
  return [nbr, ndre, msi, gndvi, savi, nddi];
}
"""

# =============================================================================
# DOWNLOAD
# =============================================================================

def download_scene(aoi_geojson, scene_date, headers, out_path):
    url     = "https://sh.dataspace.copernicus.eu/api/v1/process"
    payload = {
        "input": {
            "bounds": {"geometry": aoi_geojson},
            "data": [{"type": "sentinel-2-l2a", "dataFilter": {
                "timeRange": {"from": scene_date + "T00:00:00Z",
                              "to":   scene_date + "T23:59:59Z"},
                "maxCloudCoverage": MAX_CLOUD}}],
        },
        "output": {"width": IMG_SIZE, "height": IMG_SIZE,
                   "responses": [{"identifier": "default",
                                  "format": {"type": "image/tiff"}}]},
        "evalscript": EVALSCRIPT,
    }
    try:
        resp = request_with_retry("post", url, json=payload, headers=headers)
    except Exception as e:
        return False, str(e)
    if resp.status_code != 200:
        return False, f"HTTP {resp.status_code}: {resp.text[:300]}"
    ct = resp.headers.get("Content-Type", "")
    if "tiff" not in ct and "octet" not in ct:
        return False, f"Content-Type: {ct}"
    if len(resp.content) < 2000:
        return False, f"Muito pequeno ({len(resp.content)} bytes)"
    out_path.write_bytes(resp.content)
    return True, f"{len(resp.content)/1024:.1f} KB"

# =============================================================================
# ESTATÍSTICAS
# =============================================================================

def compute_stats(tif_path, case_id, scene_date, complaint):
    rows = []
    try:
        with rasterio.open(tif_path) as src:
            nodata = src.nodata
            for band_idx, idx_name in enumerate(INDEX_NAMES, start=1):
                raw  = src.read(band_idx).astype(np.float32)
                mask = (raw == nodata) if nodata is not None else ~np.isfinite(raw)
                data = raw[~mask]
                lo, hi = INDEX_BOUNDS.get(idx_name, (-999, 999))
                data = data[(data >= lo) & (data <= hi)]
                if data.size == 0: continue
                pcts = np.percentile(data, PERCENTILES)
                row  = {"case_id": case_id, "complaint": complaint,
                        "date": scene_date, "index": idx_name,
                        "n_pixels": int(data.size),
                        "mean": round(float(np.mean(data)), 5),
                        "std":  round(float(np.std(data)), 5)}
                for p, v in zip(PERCENTILES, pcts):
                    row[f"p{p}"] = round(float(v), 5)
                rows.append(row)
    except Exception as e:
        tqdm.write(f"    [STATS ERROR] {tif_path.name}: {e}")
    return rows

def build_timeseries(stats_list):
    if not stats_list: return pd.DataFrame()
    df   = pd.DataFrame(stats_list)
    dfs  = []
    for col in ["mean", "std", "p25", "p50", "p75"]:
        piv = df.pivot_table(index=["case_id", "complaint", "date"],
                             columns="index", values=col, aggfunc="mean")
        piv.columns = [f"{i}_{col}" for i in piv.columns]
        dfs.append(piv)
    ts = pd.concat(dfs, axis=1).reset_index()
    ts["date"] = pd.to_datetime(ts["date"])
    return ts.sort_values(["case_id", "date"])

# =============================================================================
# MAIN
# =============================================================================

def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    cases_df   = read_all_records(EXCEL_PATH)
    if cases_df.empty: return
    token      = get_token()
    headers    = get_headers(token)
    token_time = time.time()
    all_stats, case_summary = [], []

    for _, row in tqdm(cases_df.iterrows(), total=len(cases_df), desc="Casos"):
        case_id   = row["case_id"]
        complaint = row["complaint"]
        if time.time() - token_time > 1500:
            try: token = get_token(); headers = get_headers(token); token_time = time.time()
            except Exception as e: tqdm.write(f"  [TOKEN ERROR] {e}")

        case_dir = OUTPUT_DIR / f"case_{case_id}_{complaint.replace('/', '_')}"
        case_dir.mkdir(exist_ok=True)
        aoi_json, bbox = geom_to_api(row["geometry"])
        scenes  = search_scenes(bbox, str(row["start_date"]), str(row["end_date"]),
                                headers, MAX_CLOUD, MAX_SCENES)
        n_ok, n_fail, case_stats, case_dates = 0, 0, [], []

        for scene in scenes:
            scene_date = scene["properties"]["datetime"][:10]
            out_path   = case_dir / f"{case_id}_{scene_date}.tif"
            if out_path.exists() and out_path.stat().st_size > 2000:
                s = compute_stats(out_path, case_id, scene_date, complaint)
                case_stats.extend(s); all_stats.extend(s); n_ok += 1
                case_dates.append(scene_date); continue
            ok, msg = download_scene(aoi_json, scene_date, headers, out_path)
            if ok:
                s = compute_stats(out_path, case_id, scene_date, complaint)
                case_stats.extend(s); all_stats.extend(s); n_ok += 1
                case_dates.append(scene_date)
            else:
                n_fail += 1; tqdm.write(f"  [FAIL] {case_id} {scene_date}: {msg[:120]}")
            time.sleep(0.8)

        if case_stats:
            pd.DataFrame(case_stats).to_csv(case_dir / f"{case_id}_stats.csv", index=False)
            ts = build_timeseries(case_stats)
            if not ts.empty:
                ts.to_csv(case_dir / f"{case_id}_timeseries.csv", index=False)

        tqdm.write(f"  case={case_id} [{complaint}] | ok={n_ok} fail={n_fail}")
        case_summary.append({"case_id": case_id, "complaint": complaint,
                              "sheet": row["sheet"], "start": row["start_date"],
                              "end": row["end_date"], "n_scenes": n_ok,
                              "n_failed": n_fail, "dates": "; ".join(case_dates)})

    if all_stats:
        df_all = pd.DataFrame(all_stats)
        df_all.to_csv(OUTPUT_DIR / "stats_todos_indices.csv", index=False)
        ts_global = build_timeseries(all_stats)
        if not ts_global.empty:
            ts_global.to_csv(OUTPUT_DIR / "timeseries_global.csv", index=False)
            mean_cols = [c for c in ts_global.columns if c.endswith("_mean")]
            if mean_cols:
                ts_global.groupby("complaint")[mean_cols].mean().round(4).to_csv(
                    OUTPUT_DIR / "media_por_complaint.csv")
    pd.DataFrame(case_summary).to_csv(OUTPUT_DIR / "resumo_geral.csv", index=False)
    print(f"\n[OK] Concluído. Saída em {OUTPUT_DIR}")

if __name__ == "__main__":
    main()
